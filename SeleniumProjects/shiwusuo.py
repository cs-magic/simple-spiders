from time import sleep
import requests
import json
import openpyxl
from openpyxl import load_workbook,Workbook
start ='2016-01-01'
end = '2021-12-31'
pdf_list = []
# wb=load_workbook('D:/学霸/毕业论文/大论文/数据/优秀共产党员/公司名称.xlsx')
# ws=wb.active
# for cell in ws['A']:
#     keyword=cell.value
#     print(keyword)
keyword='000002'
params1 = {
    'keyWord': keyword,
    'maxNum': 10
}
r = requests.post('http://www.cninfo.com.cn/new/information/topSearch/query', params=params1)

print(r.url)

pdf_list=[]


for eachi in r.json():
    if eachi['category'] == 'A股':
        code=eachi['code']
        orgid=eachi['orgId']
        print(orgid)
        print("名称 - {} 代码 - {} ".format( eachi['zwjc'], eachi['code']))
            # 根据股票代码的头文字，判断股票交易所信息
        if code[0] == '6':
            column = 'sse'
            plate = 'sh'
        else:
            column = 'szse'
            plate = 'sz'

        # 筛选报告
        while True:
        # 设置初始页码
            page_num=1
        # 设置初始列表存储筛选结果
                

            # 设置报告筛选参数
            data = {
                'stock': '{},{}'.format(code, orgid),
                'tabName': 'fulltext',
                'pageSize': '30',
                'pageNum': str(page_num),
                'category': 'category_ndbg_szsh;',
                'seDate': '{}~{}'.format(start, end),
                'column': column,
                'plate': plate,
                'searchkey': '',
                'secid': '',
                'sortName': '',
                'sortType': '',
                'isHLtitle': ''
            }

            # 发起报告搜索请求
            r1= requests.post('http://www.cninfo.com.cn/new/hisAnnouncement/query', params=data)
            # 解析相应数据
            r1_json = r1.json()

            # 判断是否搜索失败、或者无搜索结果，如果无结果则结束
            if len(r1_json)==0:
                break



            # 遍历搜索结果，将筛选后的报告标题以及 url 以列表的形式嵌套至初识列表中
            for i in r1_json['announcements']:
                pdf_list.append([i['announcementTitle'], i['adjunctUrl']])

            # 判断是否还有下一页数据，没有的话就结束循环
            if r1_json['hasMore']=='false':
                break


            # 让页数加一，开始下一轮循环
            page_num+=1

            
        
        # 循环遍历筛选结果
        for item in pdf_list:

            # 拼接完整 url
            pdf_r = requests.get('http://static.cninfo.com.cn/' + item[1])

            # 拼接文件后缀
            file_path = 'D:\\学霸\\毕业论文\\大论文\\数据\\优秀共产党员\\年报\\'+str(code)+item[0] + '.pdf'

            # 将报告内容写入文件，保存文件
            with open(file_path, 'wb') as f:
                f.write(pdf_r.content)
            
            # 打印下载提示
            print('已下载{}'.format(item[0]))

